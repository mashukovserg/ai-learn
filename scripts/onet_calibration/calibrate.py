#!/usr/bin/env python3
"""Калибровка экспертного рейтинга профессий (модель v3.0) данными O*NET Skills.

Идея (по мотивам ILO WP140): экспертные баллы 1–10 по трём факторам сверяются с
объективным прокси, посчитанным из O*NET. Профессии с большим расхождением (|Δ|>порог)
идут на ревизию эксперта — это не автозамена, а «второе мнение».

Прокси считаются из корзин навыков (см. factor_baskets.py): средняя importance навыков
корзины → нормировка min–max по всем 894 профессиям в шкалу 0..10 → инверсия.

Вход:
  data/onet_skills.csv      — O*NET Skills (soc_code, occupation, skill, importance, level)
  expert_scores.csv         — 100 профессий с экспертными f1/f2/f3
  mapping_ru_onet.csv       — справочник profession → onet_soc (+ confidence)

Выход:
  out/calibration.csv       — по профессии: эксперт vs O*NET-прокси, Δ, флаги ревизии
  out/calibration.xlsx      — то же (если доступен openpyxl), готово вставить в Google Sheet
  out/unmapped.csv          — профессии без соответствия в O*NET

Запуск:  python3 calibrate.py
"""
import csv
import os
from collections import defaultdict

from factor_baskets import (
    FACTORS,
    ONET_METRIC,
    MISSING_SKILL_FILL,
)

HERE = os.path.dirname(os.path.abspath(__file__))
ONET_CSV = os.path.join(HERE, "data", "onet_skills.csv")
EXPERT_CSV = os.path.join(HERE, "expert_scores.csv")
MAPPING_CSV = os.path.join(HERE, "mapping_ru_onet.csv")
OUT_DIR = os.path.join(HERE, "out")

EXPERT_FIELDS = ["f1_accessibility", "f2_routineness", "f3_delegability"]

# Порог перцентиль-расхождения рангов (0..100), при котором профессия идёт на ревизию.
# Высокий порог намеренно: при слабой корреляции прокси мелкие расхождения — шум,
# инспектировать имеет смысл только крайние (обычно это ошибки маппинга или слепые зоны данных).
RANKGAP_REVIEW = 60.0
# Минимальная |Spearman|, ниже которой прокси фактора считаем непригодным и НЕ флажим по нему.
CORR_MIN = 0.20


def load_onet(path):
    """soc_code -> {skill_name: metric_value}. Пустые ячейки пропускаем."""
    by_soc = defaultdict(dict)
    soc_name = {}
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            soc = row["soc_code"]
            soc_name[soc] = row["occupation"]
            val = row.get(ONET_METRIC, "")
            if val == "":
                continue
            by_soc[soc][row["skill"]] = float(val)
    return by_soc, soc_name


def basket_score(skill_map, skills):
    """Средняя метрика навыков корзины; отсутствующий навык -> MISSING_SKILL_FILL."""
    vals = [skill_map.get(s, MISSING_SKILL_FILL) for s in skills]
    return sum(vals) / len(vals)


def compute_raw(by_soc):
    """soc -> {factor: raw_basket_score} по всем профессиям O*NET."""
    raw = {}
    for soc, skill_map in by_soc.items():
        raw[soc] = {
            fac: basket_score(skill_map, cfg["skills"])
            for fac, cfg in FACTORS.items()
        }
    return raw


def minmax_bounds(raw):
    """factor -> (min, max) по всем профессиям — база для нормировки."""
    bounds = {}
    for fac in FACTORS:
        vals = [raw[soc][fac] for soc in raw]
        bounds[fac] = (min(vals), max(vals))
    return bounds


def proxy_scores(raw_soc, bounds):
    """raw корзины одной профессии -> прокси-баллы 0..10 (с инверсией)."""
    out = {}
    for fac, cfg in FACTORS.items():
        lo, hi = bounds[fac]
        norm = 0.0 if hi == lo else (raw_soc[fac] - lo) / (hi - lo)
        score = 10.0 * (1.0 - norm) if cfg["invert"] else 10.0 * norm
        out[fac] = round(score, 1)
    return out


def load_expert(path):
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            for k in EXPERT_FIELDS:
                row[k] = float(row[k])
            rows.append(row)
    return rows


def load_mapping(path):
    m = {}
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            m[row["profession"]] = row
    return m


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    by_soc, soc_name = load_onet(ONET_CSV)
    raw = compute_raw(by_soc)
    bounds = minmax_bounds(raw)
    expert = load_expert(EXPERT_CSV)
    mapping = load_mapping(MAPPING_CSV)

    results = []
    unmapped = []
    for e in expert:
        prof = e["profession"]
        mp = mapping.get(prof)
        if not mp or not mp.get("onet_soc"):
            unmapped.append({
                "profession": prof,
                "reason": (mp or {}).get("note", "нет строки в mapping_ru_onet.csv"),
            })
            continue
        soc = mp["onet_soc"]
        if soc not in raw:
            unmapped.append({
                "profession": prof,
                "reason": f"SOC {soc} отсутствует в O*NET-файле",
            })
            continue

        px = proxy_scores(raw[soc], bounds)
        row = {
            "profession": prof,
            "onet_soc": soc,
            "onet_occupation": mp.get("onet_occupation", soc_name.get(soc, "")),
            "map_confidence": mp.get("confidence", ""),
        }
        for fac in EXPERT_FIELDS:
            row[f"expert_{fac}"] = e[fac]
            row[f"onet_{fac}"] = px[fac]
        row["expert_index"] = round(sum(e[f] for f in EXPERT_FIELDS) / 3, 1)
        row["onet_index"] = round(sum(px[f] for f in EXPERT_FIELDS) / 3, 1)
        results.append(row)

    # --- Калибровка на РАНГАХ, а не на абсолютных баллах ---------------------
    # Экспертная и O*NET-шкалы по-разному заякорены (разные нули и разброс), поэтому
    # сравнивать абсолютные значения некорректно — получаем систематический сдвиг.
    # Считаем перцентиль-ранг каждой профессии внутри замапленного набора отдельно
    # для эксперта и для O*NET, и меряем расхождение рангов (0..100).
    for fac in EXPERT_FIELDS:
        _add_percentiles(results, f"expert_{fac}", f"expert_{fac}_pct")
        _add_percentiles(results, f"onet_{fac}", f"onet_{fac}_pct")
    correlations = {
        fac: _spearman(
            [r[f"expert_{fac}"] for r in results],
            [r[f"onet_{fac}"] for r in results],
        )
        for fac in EXPERT_FIELDS
    }
    for r in results:
        review_flags = []
        for fac in EXPERT_FIELDS:
            gap = round(r[f"onet_{fac}_pct"] - r[f"expert_{fac}_pct"], 0)
            r[f"rankgap_{fac}"] = gap
            # Флаг = сильное рассогласование рангов на факторе, чей прокси не мусорный.
            if abs(correlations[fac]) >= CORR_MIN and abs(gap) >= RANKGAP_REVIEW:
                review_flags.append(fac)
        r["review_flags"] = ";".join(review_flags)
        r["needs_review"] = "ДА" if review_flags else ""

    # порядок колонок
    cols = [
        "profession", "onet_soc", "onet_occupation", "map_confidence",
        "expert_f1_accessibility", "onet_f1_accessibility", "rankgap_f1_accessibility",
        "expert_f2_routineness", "onet_f2_routineness", "rankgap_f2_routineness",
        "expert_f3_delegability", "onet_f3_delegability", "rankgap_f3_delegability",
        "expert_index", "onet_index", "review_flags", "needs_review",
    ]
    results.sort(key=lambda r: (0 if r["needs_review"] else 1, r["profession"]))

    out_csv = os.path.join(OUT_DIR, "calibration.csv")
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(results)

    out_unmapped = os.path.join(OUT_DIR, "unmapped.csv")
    with open(out_unmapped, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["profession", "reason"])
        w.writeheader()
        w.writerows(unmapped)

    out_corr = os.path.join(OUT_DIR, "correlations.csv")
    with open(out_corr, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["factor", "spearman_expert_vs_onet", "strength"])
        for fac in EXPERT_FIELDS:
            w.writerow([fac, round(correlations[fac], 2), _strength(correlations[fac])])

    _maybe_xlsx(results, cols, unmapped)
    _summary(results, unmapped, correlations)


def _add_percentiles(rows, src_key, dst_key):
    """Перцентиль-ранг (0..100) по src_key внутри набора; средний ранг при ties."""
    n = len(rows)
    order = sorted(range(n), key=lambda i: rows[i][src_key])
    i = 0
    while i < n:
        j = i
        while j + 1 < n and rows[order[j + 1]][src_key] == rows[order[i]][src_key]:
            j += 1
        avg_rank = (i + j) / 2  # 0-based средний ранг
        pct = 100.0 * avg_rank / (n - 1) if n > 1 else 50.0
        for k in range(i, j + 1):
            rows[order[k]][dst_key] = round(pct, 1)
        i = j + 1


def _spearman(xs, ys):
    n = len(xs)
    if n < 3:
        return 0.0

    def ranks(v):
        order = sorted(range(n), key=lambda i: v[i])
        r = [0.0] * n
        i = 0
        while i < n:
            j = i
            while j + 1 < n and v[order[j + 1]] == v[order[i]]:
                j += 1
            avg = (i + j) / 2 + 1
            for k in range(i, j + 1):
                r[order[k]] = avg
            i = j + 1
        return r

    rx, ry = ranks(xs), ranks(ys)
    d2 = sum((rx[i] - ry[i]) ** 2 for i in range(n))
    return 1 - 6 * d2 / (n * (n * n - 1))


def _strength(c):
    a = abs(c)
    if a >= 0.6:
        return "сильная"
    if a >= 0.4:
        return "умеренная"
    if a >= 0.2:
        return "слабая"
    return "нулевая"


def _maybe_xlsx(results, cols, unmapped):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("[i] openpyxl не установлен — xlsx пропущен, есть только CSV.")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ONET-калибровка"
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    for r in results:
        ws.append([r.get(c, "") for c in cols])
        if r["needs_review"] == "ДА":
            for c in ws[ws.max_row]:
                c.fill = flag_fill
    ws2 = wb.create_sheet("Без соответствия")
    ws2.append(["profession", "reason"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for u in unmapped:
        ws2.append([u["profession"], u["reason"]])
    out_xlsx = os.path.join(OUT_DIR, "calibration.xlsx")
    wb.save(out_xlsx)
    print(f"[✓] xlsx: {out_xlsx}")


def _summary(results, unmapped, correlations):
    n = len(results)
    flagged = [r for r in results if r["needs_review"] == "ДА"]
    print("\n=== Диагностика прокси (Spearman: экспертный балл vs O*NET Skills) ===")
    for fac in EXPERT_FIELDS:
        c = correlations[fac]
        print(f"  {fac:18s}  ρ={c:+.2f}  ({_strength(c)} связь)")
    usable = [f for f in EXPERT_FIELDS if abs(correlations[f]) >= CORR_MIN]
    print(f"\nВывод: O*NET Skills — слабый прокси. Пригодных факторов (|ρ|>={CORR_MIN}): "
          f"{len(usable)}/3. Для корректной калибровки нужны файлы O*NET Work Context и Abilities "
          f"(см. docs/onet-calibration-memo.md).")
    print(f"\nЗамаплено и посчитано: {n}")
    print(f"Без соответствия:      {len(unmapped)} ({', '.join(u['profession'] for u in unmapped)})")
    print(f"Ранговое расхождение >{RANKGAP_REVIEW} пп на пригодном факторе: {len(flagged)} проф.")


if __name__ == "__main__":
    main()
